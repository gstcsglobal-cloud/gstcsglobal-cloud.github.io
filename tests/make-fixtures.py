#!/usr/bin/env python3
"""검증용 픽스처를 xlsx에서 만든다.

만들어진 CSV에는 작업자 실명·고객사·설비호기가 들어 있다.
저장소가 공개이므로 .gitignore가 이 산출물을 막아둔다 — 절대 커밋하지 말 것.

사용법:
    python3 tests/make-fixtures.py <구글시트를_내려받은.xlsx>

구글시트에서 xlsx 받는 법: 파일 > 다운로드 > Microsoft Excel(.xlsx)
"""
import csv
import datetime
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl이 필요합니다:  pip install openpyxl")

# 시트 이름 → 산출 파일. 시트 탭 이름이 바뀌면 여기도 고친다.
SHEETS = {
    "수선실적":        "csv_wk.csv",
    "자재실적":        "csv_mat.csv",
    "설치현황":        "csv_inst.csv",
    "CIP현황 (F11)":  "csv_cip11.csv",
    "CIP현황 (F16)":  "csv_cip16.csv",
}
# 헤더만 뽑아 쓰는 것 (매퍼 단위 테스트용)
HEADER_JSON = {
    "hdr.json":     [("수선실적", "wk", 4), ("자재실적", "mat", 4), ("설치현황", "inst", 4)],
    "hdr-edu.json": [("교육현황", None, 3)],
}
HERE = os.path.dirname(os.path.abspath(__file__))


def cell(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def main(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        print(f"⚠ 시트를 찾지 못했습니다: {missing}")
        print(f"  파일에 있는 시트: {wb.sheetnames}")

    for sheet, out in SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        n = 0
        with open(os.path.join(HERE, out), "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in wb[sheet].iter_rows(values_only=True):
                w.writerow([cell(v) for v in row])
                n += 1
        print(f"  {out:<16} {n:>6}행")

    for out, specs in HEADER_JSON.items():
        data = {} if len(specs) > 1 else None
        for sheet, key, rows in specs:
            if sheet not in wb.sheetnames:
                continue
            got = []
            for i, row in enumerate(wb[sheet].iter_rows(max_row=rows, values_only=True)):
                got.append([cell(v) for v in row])
            if key:
                data[key] = got
            else:
                data = got
        if data:
            json.dump(data, open(os.path.join(HERE, out), "w", encoding="utf-8"),
                      ensure_ascii=False)
            print(f"  {out:<16} 저장")

    print("\n완료. 이 파일들은 .gitignore로 막혀 있습니다 — 커밋하지 마세요.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    main(sys.argv[1])
